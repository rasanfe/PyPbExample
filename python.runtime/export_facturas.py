# -*- coding: utf-8 -*-
"""
export_facturas.py  -  Exportacion ESTILIZADA de facturas a .xlsx con openpyxl.
Llamado desde PowerBuilder via PyPb.

  export(json_str, out_path, cols_csv="")   -> nº de filas
  export_file(in_path, out_path, cols_csv="") -> nº de filas (lee el JSON de fichero)

cols_csv: lista CSV de nombres de columna (en orden) que PB obtiene de las
columnas VISIBLES del DataWindow. Si viene vacia, se exportan todas por defecto.
Asi el Excel respeta lo que se ve en el grid (ocultar / reordenar columnas).

Demuestra lo que el SaveAs nativo de PB NO hace: cabecera con color, filas
alternas (zebra), formato de moneda en euros, fila de totales en negrita,
ancho de columnas, inmovilizar cabecera y autofiltro.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Metadatos por columna:  nombre -> (titulo, ancho, alineacion, es_moneda)
META = {
    "fecha":         ("Fecha",          12, "center", False),
    "serie":         ("Serie",           6, "center", False),
    "factura":       ("Factura",        10, "center", False),
    "cliente":       ("Cliente",         9, "center", False),
    "razon":         ("Razon social",   34, "left",   False),
    "cod_fp":        ("Cod.FP",          8, "center", False),
    "forma_pago":    ("Forma de pago",  26, "left",   False),
    "subtotal":      ("Subtotal",       14, "right",  True),
    "total_iva":     ("IVA",            12, "right",  True),
    "importe":       ("Importe",        14, "right",  True),
    "situacion":     ("Sit.",            6, "center", False),
    "fecha_factura": ("Fecha factura",  12, "center", False),
    "obra":          ("Obra",            7, "center", False),
    "descripcion":   ("Descripcion",    22, "left",   False),
    "empresa":       ("Empresa",         8, "center", False),
    "anyo":          ("Anyo",            7, "center", False),
}

DEFAULT_ORDER = ["fecha", "serie", "factura", "cliente", "razon", "cod_fp",
                 "forma_pago", "subtotal", "total_iva", "importe", "situacion",
                 "obra", "descripcion", "empresa", "anyo"]

MONEY_FMT   = '#,##0.00\\ "€"'                          # 1.234,56 €
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")   # azul oscuro
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="DDEBF7")   # azul claro (zebra)
TOTAL_FILL  = PatternFill("solid", fgColor="BDD7EE")
TOTAL_FONT  = Font(bold=True)
_thin       = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _cols(cols_csv):
    """[(name, title, width, align, money)] segun cols_csv (CSV en orden);
    si viene vacio usa el orden por defecto."""
    if cols_csv and cols_csv.strip():
        names = [c.strip() for c in cols_csv.split(",") if c.strip()]
    else:
        names = DEFAULT_ORDER
    out = []
    for name in names:
        title, width, align, money = META.get(name, (name, 14, "left", False))
        out.append((name, title, width, align, money))
    return out


def export(json_str, out_path, cols_csv=""):
    data = json.loads(json_str)
    if isinstance(data, dict):            # PB envuelve en {"data":[...]}
        data = data.get("data", [])

    cols = _cols(cols_csv)

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas 2026"

    # --- Cabecera ---
    for ci, (key, title, width, align, money) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=ci, value=title)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 20

    # --- Filas ---
    tot = {}
    r = 2
    for rec in data:
        zebra = (r % 2 == 0)
        for ci, (key, title, width, align, money) in enumerate(cols, start=1):
            v = rec.get(key, "")
            if money:
                v = _num(v)
                tot[key] = tot.get(key, 0.0) + v
            elif key in ("fecha", "fecha_factura"):
                v = str(v)[:10]
            c = ws.cell(row=r, column=ci, value=v)
            c.alignment = Alignment(horizontal=align)
            c.border = BORDER
            if money:
                c.number_format = MONEY_FMT
            if zebra:
                c.fill = ALT_FILL
        r += 1

    # --- Fila de TOTALES ---
    for ci in range(1, len(cols) + 1):
        cc = ws.cell(row=r, column=ci)
        cc.fill = TOTAL_FILL
        cc.border = BORDER
    ws.cell(row=r, column=1, value="TOTALES").font = TOTAL_FONT
    for ci, (key, title, width, align, money) in enumerate(cols, start=1):
        if money:
            cc = ws.cell(row=r, column=ci, value=round(tot.get(key, 0.0), 2))
            cc.font = TOTAL_FONT
            cc.number_format = MONEY_FMT
            cc.alignment = Alignment(horizontal="right")

    # --- Remates ---
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(cols)), r - 1)

    wb.save(out_path)
    return len(data)


def export_file(in_path, out_path, cols_csv=""):
    """Lee el JSON de un fichero (UTF-8) y delega en export()."""
    with open(in_path, encoding="utf-8") as f:
        return export(f.read(), out_path, cols_csv)
